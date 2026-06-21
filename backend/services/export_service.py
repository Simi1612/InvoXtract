import csv
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def to_json(data: dict) -> bytes:
    return json.dumps(data, indent=2, default=str).encode()


def to_csv(data: dict) -> bytes:
    invoices = data.get("invoices", [])
    buf = io.StringIO()
    if not invoices:
        return b""

    # Flatten: one row per line item, or one row per invoice if no line items
    rows = []
    for inv in invoices:
        line_items = inv.get("line_items") or []
        base = {k: v for k, v in inv.items() if k != "line_items"}
        if line_items:
            for item in line_items:
                rows.append({**base, **{f"item_{k}": v for k, v in item.items()}})
        else:
            rows.append(base)

    fieldnames = list(rows[0].keys()) if rows else []
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode()


def to_excel(data: dict) -> bytes:
    invoices = data.get("invoices", [])
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    summary_cols = ["invoice_id", "file_name", "vendor_name", "invoice_number", "date", "due_date", "subtotal", "tax", "total", "currency"]
    for col, name in enumerate(summary_cols, 1):
        cell = ws.cell(row=1, column=col, value=name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    for row, inv in enumerate(invoices, 2):
        for col, key in enumerate(summary_cols, 1):
            ws.cell(row=row, column=col, value=inv.get(key))

    # --- Line Items sheet ---
    ws2 = wb.create_sheet("Line Items")
    li_cols = ["invoice_id", "vendor_name", "description", "quantity", "unit_price", "amount"]
    for col, name in enumerate(li_cols, 1):
        cell = ws2.cell(row=1, column=col, value=name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        ws2.column_dimensions[cell.column_letter].width = 20

    row = 2
    for inv in invoices:
        for item in (inv.get("line_items") or []):
            ws2.cell(row=row, column=1, value=inv.get("invoice_id"))
            ws2.cell(row=row, column=2, value=inv.get("vendor_name"))
            ws2.cell(row=row, column=3, value=item.get("description"))
            ws2.cell(row=row, column=4, value=item.get("quantity"))
            ws2.cell(row=row, column=5, value=item.get("unit_price"))
            ws2.cell(row=row, column=6, value=item.get("amount"))
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(data: dict) -> bytes:
    invoices = data.get("invoices", [])
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Invoice Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    for inv in invoices:
        elements.append(Paragraph(f"Vendor: {inv.get('vendor_name', 'N/A')}  |  Invoice #: {inv.get('invoice_number', 'N/A')}  |  Date: {inv.get('date', 'N/A')}", styles["Normal"]))
        elements.append(Spacer(1, 6))

        line_items = inv.get("line_items") or []
        if line_items:
            table_data = [["Description", "Qty", "Unit Price", "Amount"]]
            for item in line_items:
                table_data.append([
                    item.get("description", ""),
                    item.get("quantity", ""),
                    item.get("unit_price", ""),
                    item.get("amount", ""),
                ])
            table_data.append(["", "", "Subtotal", inv.get("subtotal", "")])
            table_data.append(["", "", "Tax", inv.get("tax", "")])
            table_data.append(["", "", "Total", inv.get("total", "")])

            t = Table(table_data, colWidths=[220, 60, 80, 80])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, -3), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EFF6FF")),
            ]))
            elements.append(t)
        elements.append(Spacer(1, 20))

    doc.build(elements)
    return buf.getvalue()
